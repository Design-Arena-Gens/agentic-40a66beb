"""Utilities for flattening OpenAPI payloads and exporting them to XLSX."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


STANDARD_HEADERS = [
    "id",
    "company_name",
    "vat_code",
    "tax_code",
    "activity_status",
    "cciaa",
    "rea_code",
    "start_date",
    "registration_date",
    "address",
    "street_number",
    "toponym",
    "town",
    "province",
    "zip_code",
    "town_code",
    "region",
    "latitude",
    "longitude",
    "ateco_code",
    "ateco_description",
    "ateco_secondary",
    "ateco2022_code",
    "ateco2022_description",
    "pec",
    "sdi_code",
    "sdi_code_updated_at",
    "last_update_timestamp",
    "creation_timestamp",
    "turnover",
    "turnover_year",
    "turnover_range",
    "share_capital",
    "net_worth",
    "enterprise_size",
    "employees",
    "employee_range",
    "employee_trend",
    "website",
    "email",
    "telephone",
    "fax",
]


def flatten_company_record(record: Dict[str, Any]) -> Dict[str, Any]:
    """Transform a raw OpenAPI company payload into a flat mapping.

    The OpenAPI response nests address, classification and balance sheet data.
    This helper extracts the most relevant fields for business development
    workflows and produces a dictionary ready for tabular export.
    """

    address = (record.get("address") or {}).get("registeredOffice") or {}
    ateco = (record.get("atecoClassification") or {}).get("ateco") or {}
    ateco_secondary = (record.get("atecoClassification") or {}).get("secondaryAteco")
    ateco2022 = (record.get("atecoClassification") or {}).get("ateco2022") or {}
    ecofin = (record.get("balanceSheets") or {}).get("last") or (record.get("ecofin") or {})
    employees_block = record.get("employees") or {}

    flattened: Dict[str, Any] = {
        "id": record.get("id"),
        "company_name": record.get("companyName"),
        "vat_code": record.get("vatCode"),
        "tax_code": record.get("taxCode"),
        "activity_status": record.get("activityStatus"),
        "cciaa": record.get("cciaa"),
        "rea_code": record.get("reaCode"),
        "start_date": record.get("startDate"),
        "registration_date": record.get("registrationDate"),
        "address": address.get("streetName"),
        "street_number": address.get("streetNumber"),
        "toponym": address.get("toponym"),
        "town": address.get("town"),
        "province": address.get("province"),
        "zip_code": address.get("zipCode"),
        "town_code": address.get("townCode"),
        "region": _nested_get(address, "region", "description"),
        "latitude": _safe_coordinate(address, 1),
        "longitude": _safe_coordinate(address, 0),
        "ateco_code": ateco.get("code"),
        "ateco_description": ateco.get("description"),
        "ateco_secondary": ateco_secondary,
        "ateco2022_code": ateco2022.get("code"),
        "ateco2022_description": ateco2022.get("description"),
        "pec": record.get("pec"),
        "sdi_code": record.get("sdiCode"),
        "sdi_code_updated_at": _timestamp_to_iso(record.get("sdiCodeTimestamp")),
        "last_update_timestamp": _timestamp_to_iso(record.get("lastUpdateTimestamp")),
        "creation_timestamp": _timestamp_to_iso(record.get("creationTimestamp")),
        "turnover": ecofin.get("turnover"),
        "turnover_year": ecofin.get("turnoverYear"),
        "turnover_range": _nested_get(ecofin, "turnoverRange", "description"),
        "share_capital": ecofin.get("shareCapital"),
        "net_worth": ecofin.get("netWorth"),
        "enterprise_size": _nested_get(ecofin, "enterpriseSize", "description"),
        "employees": employees_block.get("employee") or ecofin.get("employees"),
        "employee_range": _nested_get(employees_block, "employeeRange", "description"),
        "employee_trend": employees_block.get("employeeTrend"),
        "website": _nested_get(record, "webAndSocial", "website"),
        "email": _nested_get(record, "mail", "email"),
        "telephone": _nested_get(record, "contacts", "telephoneNumber"),
        "fax": _nested_get(record, "contacts", "fax"),
    }

    return flattened


def export_to_excel(records: Iterable[Dict[str, Any]], output_path: str | Path) -> int:
    """Write the provided records to an XLSX file and return the row count."""

    rows: List[Dict[str, Any]] = [flatten_company_record(record) for record in records]
    headers = STANDARD_HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = "Aziende"

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    _autosize_columns(ws, len(headers))
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return len(rows)


def _timestamp_to_iso(timestamp: Any) -> Any:
    if isinstance(timestamp, (int, float)):
        try:
            return datetime.fromtimestamp(timestamp, tz=timezone.utc).isoformat()
        except (OverflowError, OSError, ValueError):  # pragma: no cover - defensive
            return timestamp
    return timestamp


def _safe_coordinate(address: Dict[str, Any], index: int) -> Any:
    gps = address.get("gps") or {}
    coords = gps.get("coordinates")
    if isinstance(coords, (list, tuple)) and len(coords) > index:
        return coords[index]
    return None


def _nested_get(data: Dict[str, Any], *path: str) -> Any:
    current: Any = data
    for key in path:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _autosize_columns(worksheet, count: int) -> None:
    for idx in range(1, count + 1):
        column = get_column_letter(idx)
        values = [
            worksheet.cell(row=row, column=idx).value
            for row in range(1, worksheet.max_row + 1)
        ]
        str_lengths = [len(str(value)) for value in values if value is not None]
        width = max(str_lengths, default=10)
        worksheet.column_dimensions[column].width = min(width + 2, 60)


__all__ = ["flatten_company_record", "export_to_excel"]

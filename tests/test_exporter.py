from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ateco_extractor.exporter import export_to_excel, flatten_company_record


def test_flatten_company_record_extracts_expected_fields():
    record = {
        "id": "123",
        "companyName": "Example SRL",
        "vatCode": "01234567890",
        "taxCode": "01234567890",
        "activityStatus": "ATTIVA",
        "cciaa": "VR",
        "reaCode": "123456",
        "address": {
            "registeredOffice": {
                "streetName": "Via Roma 1",
                "town": "Verona",
                "province": "VR",
                "zipCode": "37100",
                "gps": {"coordinates": [10.0, 45.0]},
                "region": {"description": "Veneto"},
            }
        },
        "atecoClassification": {
            "ateco": {"code": "1071", "description": "Produzione di pane"},
            "secondaryAteco": "107200",
            "ateco2022": {"code": "1071", "description": "Pane"},
        },
        "balanceSheets": {
            "last": {
                "turnover": 1000000,
                "turnoverYear": 2023,
                "turnoverRange": {"description": "500k-1M"},
                "shareCapital": 50000,
                "netWorth": 120000,
                "enterpriseSize": {"description": "Piccola"},
            }
        },
        "employees": {
            "employee": 12,
            "employeeRange": {"description": "11-20"},
            "employeeTrend": 5.0,
        },
        "pec": "example@pec.it",
        "sdiCode": "ABCDE12",
    }

    flat = flatten_company_record(record)

    assert flat["company_name"] == "Example SRL"
    assert flat["province"] == "VR"
    assert flat["latitude"] == 45.0
    assert flat["longitude"] == 10.0
    assert flat["ateco_code"] == "1071"
    assert flat["employees"] == 12


def test_export_to_excel_writes_file(tmp_path):
    records = [
        {
            "id": "1",
            "companyName": "Alpha",
        },
        {
            "id": "2",
            "companyName": "Beta",
        },
    ]

    output = tmp_path / "companies.xlsx"
    count = export_to_excel(records, output)

    assert count == 2
    assert output.exists()

    wb = load_workbook(output)
    ws = wb.active
    # Header row + data rows
    assert ws.max_row == 3
    assert ws.cell(row=2, column=1).value == "1"
